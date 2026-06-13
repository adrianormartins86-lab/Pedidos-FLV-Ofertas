import streamlit as st
import pandas as pd
import io
import streamlit.components.v1 as components
from streamlit_gsheets import GSheetsConnection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Gestão de Pedidos - FLV Ofertas",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# CSS GLOBAL E DE IMPRESSÃO (PALETA LARANJA)
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;700&display=swap');

:root {
    --bg-main:        #0d1117;
    --bg-card:        #161b22;
    --bg-sidebar:     #0d1117;
    --orange-dark:    #3b1c0a;  /* Laranja bem escuro para fundos */
    --orange-mid:     #96420b;  /* Laranja médio para botões e bordas */
    --orange-accent:  #e65c00;  /* Laranja vibrante principal */
    --orange-bright:  #ff7b1a;  /* Laranja claro para métricas e destaques */
    --orange-glow:    rgba(230, 92, 0, 0.25);
    --text-primary:   #e6edf3;
    --text-muted:     #7d8590;
    --text-header:    #ffddcc;  /* Texto claro com tom quente */
    --border:         #21262d;
    --border-active:  var(--orange-accent);
    --row-hover:      rgba(230, 92, 0, 0.08);
    --row-selected:   rgba(230, 92, 0, 0.18);
}

.stApp, .main { background-color: var(--bg-main) !important; color: var(--text-primary) !important; }
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif !important; }
section[data-testid="stSidebar"] { background-color: var(--bg-sidebar) !important; border-right: 1px solid var(--border); }
section[data-testid="stSidebar"] * { color: var(--text-primary) !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 14px; }

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--orange-mid) 0%, var(--orange-accent) 100%) !important;
    color: #fff !important;
    border: 1px solid var(--orange-accent) !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: .3px;
    transition: all .2s ease !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 18px var(--orange-glow) !important;
    border-color: var(--orange-bright) !important;
}
.stButton > button {
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    transition: all .2s ease !important;
}
.stButton > button:hover {
    border-color: var(--orange-accent) !important;
    color: var(--orange-bright) !important;
    transform: translateY(-1px) !important;
}
.stTextInput input, .stSelectbox > div > div {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text-primary) !important;
}
.stTextInput input:focus, .stSelectbox > div > div:focus-within {
    border-color: var(--orange-accent) !important;
    box-shadow: 0 0 0 3px var(--orange-glow) !important;
}
.title-input input {
    font-weight: 700 !important;
    font-size: 16px !important;
    color: var(--orange-bright) !important;
    padding: 2px 8px !important;
    background: transparent !important;
    border: 1px dashed #21262d !important;
}
.title-input input:focus { border: 1px dashed var(--orange-accent) !important; }

[data-testid="stDataEditor"] [data-testid="glideDataEditor"] .gdg-header-cell,
[data-testid="stDataEditor"] .dvn-stack .gdg-header {
    background-color: var(--orange-dark) !important;
    color: var(--text-header) !important;
}
[data-testid="stDataEditor"] {
    border-radius: 10px !important;
    overflow: hidden;
    border: 1px solid var(--orange-mid) !important;
    box-shadow: 0 4px 20px rgba(0,0,0,.4);
    font-size: 12px !important; 
}
[data-testid="stDataEditor"] .gdg-cell.gdg-selected,
[data-testid="stDataEditor"] .gdg-cell[data-state="focused"],
[data-testid="stDataEditor"] .gdg-cell[aria-selected="true"] {
    background-color: var(--row-selected) !important;
    outline: 2px solid var(--orange-accent) !important;
    outline-offset: -2px;
}
[data-testid="stDataEditor"] .gdg-row:hover .gdg-cell { background-color: var(--row-hover) !important; }

div[data-testid="stVerticalBlockBorderWrapper"] {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    transition: box-shadow .25s ease, border-color .25s ease;
}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {
    border-color: var(--orange-mid) !important;
    box-shadow: 0 6px 24px rgba(0,0,0,.35) !important;
}
[data-testid="stMetric"] {
    background-color: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 10px;
}
[data-testid="stMetricValue"] { color: var(--orange-bright) !important; font-weight: 700; font-size: 1.8rem !important; }
[data-testid="stMetricLabel"] { color: var(--text-muted) !important; }

.sidebar-hidden section[data-testid="stSidebar"],
.sidebar-hidden [data-testid="collapsedControl"] { display: none !important; }
.sidebar-hidden .main .block-container { max-width: 100% !important; padding-left: 2rem !important; padding-right: 2rem !important; }

.topbar-loja {
    background: linear-gradient(90deg, var(--orange-dark) 0%, #241005 100%);
    border: 1px solid var(--orange-mid);
    border-radius: 10px;
    padding: 10px 18px;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.topbar-left { display: flex; align-items: center; gap: 12px; }
.topbar-title { font-size: 18px; font-weight: 700; color: var(--text-header); }
.topbar-sub { font-size: 11px; color: var(--text-muted); margin-top: 2px; }

/* REGRAS DE IMPRESSÃO GLOBAIS - FORÇANDO PAISAGEM (LANDSCAPE) E REMOVENDO ESPAÇOS BRANCOS */
@media print {
    @page { 
        margin: 5mm 5mm; 
        size: landscape; 
    }
    .stApp, .main, body, html {
        background-color: #ffffff !important;
        background-image: none !important;
        color: #000000 !important;
        padding: 0 !important;
        margin: 0 !important;
        font-size: 11px !important; 
        -webkit-print-color-adjust: exact !important;
        print-color-adjust: exact !important;
    }
    
    /* Remove os espaços gigantes no topo do Streamlit */
    div[class^="block-container"], .block-container {
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        margin-top: 0 !important;
        max-width: 100% !important;
    }

    header, [data-testid="stSidebar"], [data-testid="stHeader"] { 
        display: none !important; 
    }
    
    [data-testid="stElementContainer"],
    [data-testid="stHorizontalBlock"],
    div[data-testid="stVerticalBlockBorderWrapper"],
    .page-header,
    hr, .stAlert, .stInfo {
        display: none !important;
    }
    
    [data-testid="stElementContainer"]:has(#print-section) {
        display: block !important;
        width: 100% !important;
    }
    
    #print-section {
        display: block !important;
        width: 100% !important;
        max-width: 100% !important;
    }
    #print-section h2 {
        font-size: 14px !important;
        margin: 0 0 5px 0 !important;
        padding-bottom: 3px !important;
        border-bottom: 1px solid #000 !important;
        color: #000 !important;
        display: block !important;
        text-align: center !important;
    }
    .print-container { 
        width: 100% !important; 
        display: flex !important;
        flex-direction: row !important;
        justify-content: space-between !important;
        align-items: flex-start !important;
        gap: 10px !important; 
    }
    
    .print-col {
        width: 49% !important; 
        flex: 0 0 49% !important;
    }

    table.print-table {
        width: 100% !important; 
        border-collapse: collapse !important;
        font-size: 10px !important; 
        color: #000000 !important;
        font-family: 'IBM Plex Sans', sans-serif !important;
        line-height: 1.1 !important;
        display: table !important;
        table-layout: fixed !important; 
    }
    table.print-table th, table.print-table td {
        border: 1px solid #000000 !important;
        padding: 2px 4px !important;
        text-align: left !important;
        color: #000000 !important;
        background-color: #ffffff !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        white-space: nowrap !important; 
    }
    
    /* Configuração de impressão agora para 6 colunas (incluindo a Média 90d) */
    table.print-table th:nth-child(1), table.print-table td:nth-child(1) { width: 10% !important; } /* Cód */
    table.print-table th:nth-child(2), table.print-table td:nth-child(2) { width: 44% !important; } /* Descrição */
    table.print-table th:nth-child(3), table.print-table td:nth-child(3) { width: 10% !important; } /* Setor */
    table.print-table th:nth-child(4), table.print-table td:nth-child(4) { width: 12% !important; } /* Est. */
    table.print-table th:nth-child(5), table.print-table td:nth-child(5) { width: 12% !important; text-align: center !important;} /* Média */
    table.print-table th:nth-child(6), table.print-table td:nth-child(6) { width: 12% !important; text-align: center !important;} /* Ped. */

    table.print-table th {
        background-color: #e0e0e0 !important;
        font-weight: bold !important;
        text-align: center !important;
    }
    table.print-table tr { break-inside: avoid !important; page-break-inside: avoid !important; }
}
@media screen {
    #print-section { display: none !important; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES E DADOS INICIAIS
# ─────────────────────────────────────────────
LOJAS = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]
NOVOS_NOMES_LOJAS = ["291", "292", "293", "294", "295", "296", "297", "298"]
MAPA_LOJAS = dict(zip(LOJAS, NOVOS_NOMES_LOJAS))
FORNECEDORES_ESPECIAIS_LINHA = ["BANANA SANTOME", "MELANCIA CARLIN", "MELANCIA MARCINHO", "RODRIGO BATATA"]

produtos_iniciais = [
    {"Cód.Prime": None, "Código": 1571, "Descrição": "Abacate Cx 20 Kg", "Tipo": "Box"},
    {"Cód.Prime": None, "Código": 2614, "Descrição": "Abacaxi Doce Mel Cx c/7", "Tipo": "Box"},
    # ... os demais continuam iguais ...
    {"Cód.Prime": None, "Código": 67, "Descrição": "Vagem kg Cx 11kg", "Tipo": "Pedra"}
]

mapa_inicial_codigos = {
    "NIDE": [45, 49, 67, 57, 46, 48, 47], "Claudir Mendes": [57, 46, 49], "SANDRO": [75],
    "DENIZE": [45, 49, 67, 57, 46, 48, 47, 41, 52, 69, 56], "JOVANO": [1746, 88, 49, 140, 85],
    "JEFINHO": [85, 140, 256, 267, 88, 57, 45, 49, 1662, 46], "LUCIANO": [61, 41, 49, 56, 45],
    "THIAGO": [61, 91, 67, 74, 49, 52, 45, 56], "CRISTIAN": [40, 949, 42, 83, 68, 538, 78],
    "ROGERIO NARANTE": [538], "FERNANDO NARANTE": [46], "SILVIO MAND SALSA": [76],
    "HORTA": [108, 109], "GLAUCIA MACIEL": [84, 85], "ALEMÃO": [39], "RENAN SS": [72],
    "NEGUIN": [85, 86, 88, 61, 45], "RODRIGO CHANAN": [85, 86, 88, 61, 1662, 140],
    "MARCELO MORANGO": [58], "JOÃO BATISTA": [79, 60, 56, 1662, 69], "GIACOMELLO": [95],
    "PRIMO": [240, 86, 49, 45, 88, 85], "RENATO MANDIOCA": [75], "THIAGO SERRA": [91, 49, 45, 56, 61],
    "TICO": [236, 237, 707, 2730, 42, 581, 78, 546, 80, 83, 540, 949, 40, 110, 68, 109],
    "ALGACIR": [1516, 53], "MAURICIO": [62], "PAULO IGASHIBAHI": [47, 48],
    "GILSOM BATATA": [508, 551], "DORI BATATA": [508, 551], "BANANA SANTOME": [2567, 2569, 2568],
    "MELANCIA CARLIN": [1], "MELANCIA MARCINHO": [673, 1, 3003], "RODRIGO BATATA": [508]
}

# ─────────────────────────────────────────────
# CONEXÃO GOOGLE SHEETS & FUNÇÕES DE DADOS
# ─────────────────────────────────────────────
conn = st.connection("gsheets", type=GSheetsConnection)

# VARIÁVEIS DAS ABAS (OFERTAS)
WS_PRODUTOS = "Ofertas_Produtos"
WS_PEDIDOS  = "Ofertas_Pedidos"
WS_MEDIA_90 = "Ofertas_90d"  # <--- NOVA ABA ADICIONADA AQUI

@st.cache_data(ttl=15)
def carregar_banco():
    df_prod = conn.read(worksheet=WS_PRODUTOS)
    df_ped = conn.read(worksheet=WS_PEDIDOS)
    
    # Carregamento seguro da nova aba de Médias
    try:
        df_media_90 = conn.read(worksheet=WS_MEDIA_90)
        if df_media_90.empty:
            df_media_90 = pd.DataFrame(columns=["loja", "codigo", "qtde"])
    except Exception:
        df_media_90 = pd.DataFrame(columns=["loja", "codigo", "qtde"])

    mudou_algo = False
    
    # NORMALIZAÇÃO BLINDADA DE COLUNAS
    df_prod.columns = df_prod.columns.astype(str).str.strip()
    df_ped.columns = df_ped.columns.astype(str).str.strip()
    
    renomear_prod = {}
    for c in df_prod.columns:
        cl = c.lower()
        if "prime" in cl: 
            renomear_prod[c] = "Cód.Prime"
        elif cl in ["código", "codigo", "cod", "cód", "cod.", "cód.", "cód. iceasa", "cód.interno"]: 
            renomear_prod[c] = "Código"
        elif cl in ["descrição", "descricao", "desc"]: 
            renomear_prod[c] = "Descrição"
        elif cl in ["tipo", "setor"]: 
            renomear_prod[c] = "Tipo"
    df_prod = df_prod.rename(columns=renomear_prod)

    renomear_ped = {}
    for c in df_ped.columns:
        cl = c.lower()
        if cl in ["código", "codigo", "cod", "cód", "cod.", "cód.", "cód. iceasa", "cód.interno"]: 
            renomear_ped[c] = "Código"
        elif cl in ["r$preço", "r$ preço", "preço", "preco"]: 
            renomear_ped[c] = "R$Preço"
        elif cl in ["obs:", "obs", "observacao", "observação"]: 
            renomear_ped[c] = "OBS:"
    df_ped = df_ped.rename(columns=renomear_ped)

    # Força Bruta para garantir colunas obrigatórias em Produtos
    for col in ["Cód.Prime", "Código", "Descrição", "Tipo"]:
        if col not in df_prod.columns:
            df_prod[col] = None if col == "Cód.Prime" else ""
            mudou_algo = True
            
    if df_prod.empty:
        df_prod = pd.DataFrame(produtos_iniciais)
        for loja in LOJAS: df_prod[loja] = True
        conn.update(worksheet=WS_PRODUTOS, data=df_prod)
        mudou_algo = True

    # Força Bruta para garantir colunas obrigatórias em Pedidos
    for col in ["Código", "R$Preço", "OBS:"] + LOJAS:
        if col not in df_ped.columns:
            if col == "R$Preço": df_ped[col] = 0.0
            elif col == "OBS:": df_ped[col] = ""
            else: df_ped[col] = 0
            mudou_algo = True
            
    if df_ped.empty:
        df_ped["Código"] = df_prod["Código"]
        conn.update(worksheet=WS_PEDIDOS, data=df_ped)
        mudou_algo = True

    # Remove Cód.Prime do Pedidos se o usuário criou lá no Sheets sem querer
    for drop_col in ["Cód.Prime", "Prime"]:
        if drop_col in df_ped.columns:
            df_ped = df_ped.drop(columns=[drop_col])

    # Adiciona produtos novos nos Pedidos
    novos_ped = df_prod[~df_prod["Código"].isin(df_ped["Código"])]["Código"]
    if not novos_ped.empty:
        df_n_ped = pd.DataFrame({"Código": novos_ped})
        for col in LOJAS: df_n_ped[col] = 0
        df_n_ped["R$Preço"] = 0.0
        df_n_ped["OBS:"] = ""
        df_ped = pd.concat([df_ped, df_n_ped], ignore_index=True)
        conn.update(worksheet=WS_PEDIDOS, data=df_ped)
        mudou_algo = True

    # Converte os tipos de dados
    for loja in LOJAS:
        df_ped[loja] = pd.to_numeric(df_ped[loja], errors='coerce').fillna(0).astype(int)
        df_prod[loja] = df_prod[loja].fillna(False).astype(bool)

    df_ped["R$Preço"] = pd.to_numeric(df_ped["R$Preço"], errors='coerce').fillna(0.0)
    df_ped["OBS:"] = df_ped["OBS:"].fillna("").astype(str)
    df_prod["Cód.Prime"] = pd.to_numeric(df_prod["Cód.Prime"], errors='coerce').fillna(0).astype(int)

    if mudou_algo:
        st.cache_data.clear()

    return df_prod, df_ped, df_media_90  # Retornando a variável nova também

df_produtos, df_pedidos, df_media_90 = carregar_banco()

LISTA_NOMES_PRODUTOS = [str(x) for x in df_produtos['Descrição'].unique()]

lista_cfg = []
for f, cods in mapa_inicial_codigos.items():
    for c in cods:
        desc_match = df_produtos[df_produtos['Código'] == c]['Descrição']
        if not desc_match.empty:
            desc = desc_match.values[0]
            lista_cfg.append({"Fornecedor": f, "Produto": desc})
df_fornecedores_config = pd.DataFrame(lista_cfg)

if 'reset_counter' not in st.session_state:
    st.session_state['reset_counter'] = 0

# ─────────────────────────────────────────────
# SISTEMA DE LOGIN
# ─────────────────────────────────────────────
if 'usuario_logado' not in st.session_state:
    st.session_state['usuario_logado'] = None

if st.session_state['usuario_logado'] is None:
    st.write("<br><br>", unsafe_allow_html=True)
    _, col2, _ = st.columns([1, 1.4, 1])

    with col2:
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown("""
                    <h2 style='margin-bottom:0'>Portal de Pedidos</h2>
                    <p style='color:#7d8590;font-size:14px;margin-top:4px'>FLV Ofertas — Molicenter</p>
                """, unsafe_allow_html=True)
            with h2:
                st.write("")
                try:
                    st.image("passaro_logo.png", width=60)
                except Exception:
                    st.markdown("🐦", unsafe_allow_html=True)

            st.divider()

            usuarios_permitidos = ["Selecione..."] + ["Administrador"] + LOJAS
            usuario_selecionado = st.selectbox("👤 Usuário de acesso:", usuarios_permitidos)
            senha_digitada = st.text_input("🔑 Senha de acesso:", type="password", autocomplete="off")

            st.write("<br>", unsafe_allow_html=True)

            if st.button("Entrar no Sistema", type="primary", use_container_width=True):
                if usuario_selecionado == "Selecione...":
                    st.error("⚠️ Por favor, selecione um usuário.")
                elif usuario_selecionado == "Administrador" and senha_digitada == "moli0000":
                    st.session_state['usuario_logado'] = usuario_selecionado
                    st.rerun()
                elif usuario_selecionado in LOJAS and senha_digitada == "moli1234":
                    st.session_state['usuario_logado'] = usuario_selecionado
                    st.rerun()
                elif senha_digitada:
                    st.error("⚠️ Senha incorreta. Tente novamente.")

            st.markdown('<p style="font-size: 11px; color: #7d8590; text-align: center; margin-top: 10px;">🔒 Acesso restrito — Molicenter © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────
# PÓS-LOGIN
# ─────────────────────────────────────────────
usuario_atual = st.session_state['usuario_logado']
acesso_total  = usuario_atual == "Administrador"

if not acesso_total:
    st.markdown("""
    <script>
        document.body.classList.add('sidebar-hidden');
        const root = window.parent.document.querySelector('.stApp');
        if (root) root.classList.add('sidebar-hidden');
    </script>
    <style>
        section[data-testid="stSidebar"] { display: none !important; }
        [data-testid="collapsedControl"]  { display: none !important; }
        .main .block-container {
            max-width: 100% !important;
            padding-left: 2.5rem !important;
            padding-right: 2.5rem !important;
        }
    </style>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# MODAL ZERAR TUDO
# ─────────────────────────────────────────────
@st.dialog("🚨 Confirmação Necessária")
def modal_zerar_pedidos():
    st.markdown("Tem certeza que deseja **zerar todos os pedidos** de todas as lojas?")
    st.markdown("⚠️ *Esta ação limpará também os preços e observações direto no Google Sheets e não poderá ser desfeita.*")
    
    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, zerar tudo", type="primary", use_container_width=True):
            st.session_state['reset_counter'] += 1
            _, df_ped, _ = carregar_banco()
            
            df_ped[LOJAS] = 0
            df_ped["R$Preço"] = 0.0
            df_ped["OBS:"] = ""
            
            conn.update(worksheet=WS_PEDIDOS, data=df_ped)
            st.cache_data.clear()
            st.rerun()

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("passaro_logo.png", width=72)
    except Exception:
        st.markdown("🐦")

    st.markdown(f"### Olá, **{usuario_atual}**")
    st.caption("Sistema de Pedidos Integrado")
    st.divider()

    if acesso_total:
        perfil_navegacao = st.radio("📍 Navegação:", [
            "Separação e Fechamento",
            "Visão das Lojas",
            "Visão Fornecedores (Ademilto)",
            "Catálogo de Produtos"
        ])
    else:
        perfil_navegacao = "Visão das Lojas"

    st.divider()

    total_preenchidos = (df_pedidos[LOJAS] > 0).any(axis=1).sum()
    st.metric("Itens c/ pedido", total_preenchidos, help="Itens que têm ao menos 1 quantidade preenchida")

    st.divider()
    if st.button("🚪 Sair / Logout", use_container_width=True):
        st.session_state['usuario_logado'] = None
        st.rerun()

# ─────────────────────────────────────────────
# HELPER: gera o arquivo Excel formatado (MANTIDO IGUAL)
# ─────────────────────────────────────────────
def _gerar_excel_formatado(df_editado_admin, filtro_setor):
    HDR_BG    = "C55A11"   
    HDR_FG    = "FFFFFF"   
    GREEN_ROW = "E2EFDA"   
    WHITE_ROW = "FFFFFF"   
    TOTAL_BG  = "C6EFCE"   
    PRICE_BG  = "FCE4D6"   

    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    df_exp = df_editado_admin.copy()
    
    df_exp = df_exp.drop(columns=["Cód.Prime"], errors="ignore")
    df_exp = df_exp.rename(columns=MAPA_LOJAS)

    if filtro_setor in ("Box", "Pedra"):
        df_exp = df_exp.rename(columns={
            "Código":      "COD.ICEASA",
            "Descrição":   "PRODUTOS MOLICENTER",
            "TOTAL GERAL": "TOTAL",
            "R$Preço":     "PREÇO",
        })
    else:
        df_exp = df_exp.rename(columns={"Código": "Cód. Iceasa"})

    cod_col       = "COD.ICEASA" if filtro_setor in ("Box", "Pedra") else "Cód. Iceasa"
    prod_col      = "PRODUTOS MOLICENTER" if filtro_setor in ("Box", "Pedra") else "Descrição"
    tot_col       = "TOTAL"        if filtro_setor in ("Box", "Pedra") else "TOTAL GERAL"
    pre_col       = "PREÇO"        if filtro_setor in ("Box", "Pedra") else "R$Preço"
    obs_col       = "OBS:"

    base_cols = [cod_col, prod_col]
    if "Tipo" in df_exp.columns:
        base_cols.append("Tipo")

    store_cols  = NOVOS_NOMES_LOJAS              
    spacer_cols = [" " * i for i in range(1, 7)] 
 
    for s in spacer_cols:
        df_exp[s] = ""

    final_cols = base_cols + store_cols + spacer_cols + [tot_col, pre_col, obs_col]
    df_exp = df_exp[[c for c in final_cols if c in df_exp.columns]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos FLV"

    HEADER_ROW = 2    
    DATA_START  = 3

    cols = list(df_exp.columns)

    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=col_name)
        cell.font      = Font(bold=True, color=HDR_FG, name="Arial", size=9)
        cell.fill      = PatternFill("solid", start_color=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(cols))}{HEADER_ROW}"

    for ri, (_, row) in enumerate(df_exp.iterrows(), DATA_START):
        for ci, col_name in enumerate(cols, 1):
            raw = row[col_name]
            if raw == 0 or raw == 0.0 or str(raw).strip() in ("0", "0.0", "nan", ""):
                raw = None

            cell = ws.cell(row=ri, column=ci, value=raw)
            cell.font   = Font(name="Arial", size=9, bold=True) 
            cell.border = brd

            if col_name in (prod_col, "Descrição", "PRODUTOS MOLICENTER"):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == tot_col:
                idx_ini = get_column_letter(len(base_cols) + 1)
                idx_fim = get_column_letter(len(base_cols) + len(store_cols))
                cell.value = f'=IF(SUM({idx_ini}{ri}:{idx_fim}{ri})=0,"",SUM({idx_ini}{ri}:{idx_fim}{ri}))'
                cell.fill = PatternFill("solid", start_color=TOTAL_BG)
                cell.font = Font(name="Arial", size=9, bold=True)
                            
            elif col_name == pre_col:
                cell.fill = PatternFill("solid", start_color=PRICE_BG)
                if raw is not None:
                    cell.number_format = '[$R$-pt-BR] #,##0.00'
            
            else:
                if ci > len(base_cols) and ci <= len(base_cols) + len(store_cols):
                    col_bg = GREEN_ROW if ci % 2 != 0 else WHITE_ROW
                else:
                    col_bg = WHITE_ROW  
                
                cell.fill = PatternFill("solid", start_color=col_bg)

    widths = {
        cod_col:       9,
        prod_col:      34,
        "Tipo":        8,
        tot_col:       8,
        pre_col:       12,
        obs_col:       22,
    }
    for s in store_cols:
        widths[s] = 6
    for s in spacer_cols:
        widths[s] = 3

    for ci, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col_name, 8)

    for ci, col_name in enumerate(cols, 1):
        if col_name.strip() == "":
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    freeze_col = len(base_cols) + 1
    ws.freeze_panes = f"{get_column_letter(freeze_col)}{DATA_START}"

    ws.row_dimensions[1].height = 6   
    ws.row_dimensions[HEADER_ROW].height = 18
    for ri in range(DATA_START, DATA_START + len(df_exp)):
        ws.row_dimensions[ri].height = 15

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# ROTA 1: SEPARAÇÃO E FECHAMENTO
# ─────────────────────────────────────────────
if perfil_navegacao == "Separação e Fechamento":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--orange-dark) 0%, #241005 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">📊</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Separação e Fechamento</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Consolidado geral de quantidades (O estoque das lojas não é exibido aqui)</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        filtro_setor = st.radio("🔍 Filtrar Exibição por Setor:", ["Todos", "Box", "Pedra"], horizontal=True)
        st.write("<br>", unsafe_allow_html=True)

        df_base  = df_produtos[["Cód.Prime", "Código","Descrição","Tipo"]]
        df_final = pd.merge(df_base, df_pedidos, on="Código")
        df_final["TOTAL GERAL"] = df_final[LOJAS].sum(axis=1)
        cols_order = ["Cód.Prime", "Código", "Descrição", "Tipo"] + LOJAS + ["TOTAL GERAL", "R$Preço", "OBS:"]
        df_final = df_final[cols_order]

        if filtro_setor != "Todos":
            df_final = df_final[df_final["Tipo"] == filtro_setor].reset_index(drop=True)
            df_final = df_final.drop(columns=["Tipo"])

        # Oculta zeros no código Prime para estética
        df_final["Cód.Prime"] = df_final["Cód.Prime"].replace(0, None)

        col_cfg = {
            "Cód.Prime":   st.column_config.NumberColumn("Cód. Prime", width=70, format="%d", disabled=True),
            "Código":      st.column_config.NumberColumn("Cód. Iceasa", width=80, format="%d", disabled=True),
            "Descrição":   st.column_config.TextColumn(disabled=True),
            "TOTAL GERAL": st.column_config.NumberColumn("TOTAL ▶", width=90, format="%d", disabled=True),
            "R$Preço":     st.column_config.NumberColumn("R$ Preço", width=100, format="R$ %.2f", min_value=0.0, step=0.01),
            "OBS:":        st.column_config.TextColumn("OBS:", width=200)
        }
        
        if "Tipo" in df_final.columns:
            col_cfg["Tipo"] = st.column_config.TextColumn("Setor", width=100, disabled=True)
            
        for loja, novo_nome in MAPA_LOJAS.items():
            col_cfg[loja] = st.column_config.NumberColumn(novo_nome, format="%d", min_value=0, step=1)

        df_editado_admin = st.data_editor(
            df_final, hide_index=True, use_container_width=True,
            height=580, column_config=col_cfg,
            key=f"admin_editor_{st.session_state['reset_counter']}"
        )
        
        # Oculta o Cód.Prime na hora da impressão para manter o padrão antigo
        df_imprimir = df_editado_admin.copy()
        df_imprimir = df_imprimir.drop(columns=["Cód.Prime"], errors="ignore")
        df_imprimir = df_imprimir.rename(columns={"Código": "Cód. Iceasa"})
        html_table = df_imprimir.to_html(index=False, classes=["print-table", "print-sep"])
        st.markdown(f"""<div id="print-section">
            <h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
                Resumo de Separação — FLV Ofertas
            </h2>
            <div class="print-container">
                {html_table}
            </div>
        </div>""", unsafe_allow_html=True)

        st.divider()
        col_salvar, col_csv, col_excel, col_limpa, _ = st.columns([2.5, 1.5, 1.5, 2, 2.5])

        with col_salvar:
            if st.button("💾 Salvar Alterações", type="primary", use_container_width=True):
                _, df_ped_fresco, _ = carregar_banco()
                for _, row in df_editado_admin.iterrows():
                    mask = df_ped_fresco["Código"] == row["Código"]
                    for loja in LOJAS:
                        df_ped_fresco.loc[mask, loja] = row[loja]
                    df_ped_fresco.loc[mask, "R$Preço"] = row["R$Preço"]
                    df_ped_fresco.loc[mask, "OBS:"] = row["OBS:"]
                
                conn.update(worksheet=WS_PEDIDOS, data=df_ped_fresco)
                st.cache_data.clear()
                st.success("✅ Ajustes, preços e observações salvos com sucesso no Sheets!")
                st.rerun()

        with col_csv:
            df_csv = df_editado_admin.copy()
            df_csv = df_csv.drop(columns=["Cód.Prime"], errors="ignore")
            df_csv = df_csv.rename(columns=MAPA_LOJAS)
            csv = df_csv.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ CSV", data=csv, file_name="separacao_semanal_ofertas.csv", mime="text/csv", use_container_width=True)
            
        with col_excel:
            excel_bytes = _gerar_excel_formatado(df_editado_admin, filtro_setor)
            nome_arquivo_excel = f"separacao_ofertas_{filtro_setor.lower()}.xlsx" if filtro_setor != "Todos" else "separacao_semanal_ofertas.xlsx"
            st.download_button(
                "⬇️ Excel",
                data=excel_bytes,
                file_name=nome_arquivo_excel,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col_limpa:
            if st.button("🚨 Zerar Pedidos", use_container_width=True):
                modal_zerar_pedidos()

# ─────────────────────────────────────────────
# ROTA 2: VISÃO DAS LOJAS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão das Lojas":
    if acesso_total:
        loja_selecionada = st.selectbox("👁️ Visualizar como:", LOJAS)
    else:
        loja_selecionada = usuario_atual

    col_info, col_logout = st.columns([8, 2])
    with col_info:
        id_loja = MAPA_LOJAS.get(loja_selecionada, loja_selecionada)
        st.markdown(f"""
        <div class="topbar-loja">
            <div class="topbar-left">
                <span style="font-size:22px">📋</span>
                <div>
                    <div class="topbar-title">{loja_selecionada} ({id_loja}) — FLV Ofertas</div>
                    <div class="topbar-sub">O estoque é carregado pelo ERP Prime. Preencha a quantidade necessária para o pedido.</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_logout:
        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
        if st.button("🚪 Sair / Logout", use_container_width=True):
            st.session_state['usuario_logado'] = None
            st.rerun()

    df_visiveis = df_produtos[df_produtos[loja_selecionada] == True]
    df_loja = df_visiveis[["Cód.Prime", "Código", "Descrição", "Tipo"]].copy()
    
    mapa_banco_erp = {
        "Loja 01": "001", "Loja 02": "002", "Loja 03": "003",
        "Loja 04": "004", "Loja 05": "005", "Loja 06": "006",
        "Loja 07": "007", "Loja 08": "008"
    }
    cod_empresa_banco = mapa_banco_erp.get(loja_selecionada, "001")

    # ------------------ ESTOQUE VIA POSTGRESQL ----------------------
    try:
        conn_pg = st.connection("banco_erp", type="sql")
        
        query_erp = f"""
            SELECT cade_codempresa,
                   cade_codigo,
                   cadp_descricao,
                   estoque
            FROM "python_estoque"
            WHERE cade_codempresa::text = '{cod_empresa_banco}'
            ORDER BY cade_codempresa, cade_codigo
        """
        df_erp = conn_pg.query(query_erp, ttl=300)

        if not df_erp.empty:
            df_erp["cade_codigo"] = pd.to_numeric(df_erp["cade_codigo"], errors='coerce').fillna(0).astype(int)
            df_erp = df_erp.rename(columns={"cade_codigo": "Cód.Prime", "estoque": "Estoque"})
            df_loja["Cód.Prime"] = pd.to_numeric(df_loja["Cód.Prime"], errors='coerce').fillna(0).astype(int)
            df_loja = pd.merge(df_loja, df_erp[["Cód.Prime", "Estoque"]], on="Cód.Prime", how="left")
        else:
            df_loja["Estoque"] = 0

    except Exception as e:
        if "No database configured" in str(e) or "missing" in str(e).lower():
             st.error("⚠️ Aviso: As credenciais do banco_erp precisam estar configuradas no Streamlit Secrets.")
        else:
             st.error(f"⚠️ Erro ao puxar dados do ERP: {e}")
        df_loja["Estoque"] = 0
    # -----------------------------------------------------------------

    df_loja["Estoque"] = df_loja["Estoque"].fillna(0).astype(int)
    df_qtd = df_pedidos[["Código", loja_selecionada]].rename(columns={loja_selecionada: "Qtde"})
    df_loja = pd.merge(df_loja, df_qtd, on="Código", how="left")
    
    # --- CRUZAMENTO DA MÉDIA 90 DIAS ---------------------------------
    df_media_loja = df_media_90[df_media_90['loja'].astype(str).str.zfill(3) == cod_empresa_banco.zfill(3)].copy()
    df_media_loja = df_media_loja.rename(columns={"codigo": "Cód.Prime", "qtde": "Média 90d"})
    df_media_loja["Cód.Prime"] = pd.to_numeric(df_media_loja["Cód.Prime"], errors='coerce').fillna(0).astype(int)

    df_loja = pd.merge(df_loja, df_media_loja[["Cód.Prime", "Média 90d"]], on="Cód.Prime", how="left")
    df_loja["Média 90d"] = df_loja["Média 90d"].fillna(0.0).round(2)
    # -----------------------------------------------------------------

    df_loja["Cód.Prime"] = df_loja["Cód.Prime"].replace(0, None)
    # 👇 ADICIONE ESTA LINHA AQUI 👇
    df_loja = df_loja[["Cód.Prime", "Código", "Descrição", "Tipo", "Estoque", "Média 90d", "Qtde"]]

    with st.container(border=True):
        col_cfg_loja = {
            "Código":         None, 
            "Cód.Prime":      st.column_config.NumberColumn("Cód. Prime", width=70, format="%d", disabled=True),
            "Descrição":      st.column_config.TextColumn(width=340, disabled=True),
            "Tipo":           st.column_config.TextColumn("Setor", width=70, disabled=True),
            "Estoque":        st.column_config.NumberColumn("📦 Estoque", width=80, disabled=True),
            "Média 90d":      st.column_config.NumberColumn("📈 Média 90d", width=90, disabled=True, format="%.2f"),
            "Qtde":           st.column_config.NumberColumn("🛒 Pedido", width=90, min_value=0, step=1)
        }
        
        df_editado = st.data_editor(
            df_loja, column_config=col_cfg_loja,
            hide_index=True, use_container_width=True, height=520,
            key=f"loja_editor_{st.session_state['reset_counter']}"
        )

        df_imprimir = df_editado.copy()
        df_imprimir = df_imprimir.drop(columns=["Cód.Prime"], errors="ignore")
        df_imprimir["Código"] = df_imprimir["Código"].fillna(0).astype(int).astype(str)
        # Renomeia para impressão. Adicionado Média
        df_imprimir = df_imprimir.rename(columns={"Código": "Cód", "Tipo": "Setor", "Estoque": "Est.", "Média 90d": "Média", "Qtde": "Ped."})
        df_imprimir = df_imprimir[["Cód", "Descrição", "Setor", "Est.", "Média", "Ped."]] 
        
        meio = (len(df_imprimir) // 2) + (len(df_imprimir) % 2)
        df1 = df_imprimir.iloc[:meio]
        df2 = df_imprimir.iloc[meio:]
        
        html1 = df1.to_html(index=False, classes="print-table")
        html2 = df2.to_html(index=False, classes="print-table") if not df2.empty else ""
        
        st.markdown(f"""
        <div id="print-section">
            <h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
                Resumo do Pedido — {loja_selecionada} ({id_loja})
            </h2>
            <div class="print-container">
                <div class="print-col">{html1}</div>
                <div class="print-col">{html2}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        itens_com_pedido = int((df_editado["Qtde"] > 0).sum())
        total_itens      = len(df_editado)
        total_unidades   = int(df_editado["Qtde"].sum())
        pct              = round(itens_com_pedido / total_itens * 100) if total_itens else 0

        st.divider()
        m1, m2, m3, col_print, col_btn = st.columns([2.5, 2.2, 1.8, 1.5, 3])
        with m1: st.metric("Itens preenchidos (Pedido)", f"{itens_com_pedido} / {total_itens}")
        with m2: st.metric("Total de unidades", total_unidades)
        with m3: st.metric("Cobertura", f"{pct}%")
        
        with col_print:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("🖨️ Imprimir", use_container_width=True):
                components.html(
                    "<script>"
                    "window.parent.print();"
                    "</script>", 
                    height=0
                )
                
        with col_btn:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("💾 Salvar Pedido da Semana", type="primary", use_container_width=True):
                _, df_ped_fresco, _ = carregar_banco()
                
                for _, row in df_editado.iterrows():
                    mask_ped = df_ped_fresco["Código"] == row["Código"]
                    df_ped_fresco.loc[mask_ped, loja_selecionada] = row["Qtde"]
                
                conn.update(worksheet=WS_PEDIDOS, data=df_ped_fresco)
                st.cache_data.clear()
                st.success(f"✅ Pedido da {loja_selecionada} salvo na nuvem com sucesso!")

# ─────────────────────────────────────────────
# ROTA 3: VISÃO FORNECEDORES (ADEMILTO)
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão Fornecedores (Ademilto)":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--orange-dark) 0%, #241005 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🚚</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Visão Fornecedores (Ademilto) - Modo de Edição Livre</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Altere nomes, códigos e totais livremente antes de gerar seu print. Edite os campos diretamente na tabela.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_base_produtos = df_produtos[["Código", "Descrição"]]
    df_base_pedidos = df_pedidos.copy()
    df_base_pedidos["Total"] = df_base_pedidos[LOJAS].sum(axis=1)

    df_consolidado = pd.merge(df_base_produtos, df_base_pedidos[["Código", "Total", "R$Preço"]], on="Código", how="inner")
    nomes_fornecedores = df_fornecedores_config["Fornecedor"].unique()
    
    for i in range(0, len(nomes_fornecedores), 2):
        cols = st.columns(2, gap="small")
        for j, fornecedor in enumerate(nomes_fornecedores[i:i+2]):
            descricoes_fornecedor = df_fornecedores_config[df_fornecedores_config["Fornecedor"] == fornecedor]["Produto"].tolist()
            codigos_do_fornecedor = df_base_produtos[df_base_produtos["Descrição"].isin(descricoes_fornecedor)]["Código"].tolist()
            
            with cols[j]:
                with st.container(border=True):
                    st.markdown('<div class="title-input">', unsafe_allow_html=True)
                    st.text_input("Fornecedor", value=f"🛒 {fornecedor}", label_visibility="collapsed", key=f"title_{fornecedor}_{st.session_state['reset_counter']}")
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    if fornecedor in FORNECEDORES_ESPECIAIS_LINHA:
                        df_ped_esp = df_base_pedidos[df_base_pedidos["Código"].isin(codigos_do_fornecedor)]
                        dict_lojas = {"Visão": LOJAS + ["TOTAL"]}
                        col_configs_especial = {"Visão": st.column_config.TextColumn("Visão", disabled=False)}
                        
                        for cod in codigos_do_fornecedor:
                            desc_series = df_base_produtos[df_base_produtos["Código"] == cod]["Descrição"]
                            desc = desc_series.values[0] if not desc_series.empty else "Prod"
                            
                            partes = desc.split()
                            palavra = " ".join(partes[:2]) if len(partes) > 1 else desc
                            nome_col = f"{cod} - {palavra}"
                            
                            valores_lojas = []
                            for loja in LOJAS:
                                val = df_ped_esp[df_ped_esp["Código"] == cod][loja].values
                                valores_lojas.append(int(val[0]) if len(val) > 0 else 0)
                            
                            valores_lojas.append(sum(valores_lojas))
                            dict_lojas[nome_col] = valores_lojas
                            col_configs_especial[nome_col] = st.column_config.NumberColumn(nome_col, format="%d", disabled=False)
                            
                        df_especial = pd.DataFrame(dict_lojas)
                        altura_esp = int((len(df_especial) + 2) * 36) + 5
                        
                        st.data_editor(
                            df_especial, 
                            hide_index=True, 
                            use_container_width=True, 
                            column_config=col_configs_especial,
                            height=altura_esp,
                            num_rows="fixed",
                            key=f"forn_esp_{fornecedor}_{st.session_state['reset_counter']}"
                        )
                    
                    else:
                        df_fornecedor = df_consolidado[df_consolidado["Código"].isin(codigos_do_fornecedor)].copy()
                        df_fornecedor = df_fornecedor.rename(columns={"Código": "Cód", "Descrição": "Produtos", "R$Preço": "R$ Preço"})
                        df_fornecedor["R$ Total"] = df_fornecedor["Total"] * df_fornecedor["R$ Preço"]
                        df_exibicao = df_fornecedor[["Cód", "Produtos", "Total", "R$ Preço", "R$ Total"]].copy()

                        df_exibicao['Produtos'] = pd.Categorical(df_exibicao['Produtos'], categories=LISTA_NOMES_PRODUTOS)
                        altura_dinamica = int((len(df_exibicao) + 2) * 36) + 5
                        
                        col_cfg_forn = {
                            "Cód": st.column_config.NumberColumn(disabled=False, format="%d"),
                            "Produtos": st.column_config.SelectboxColumn("Produtos", options=LISTA_NOMES_PRODUTOS, disabled=False),
                            "Total": st.column_config.NumberColumn("Total", disabled=False, format="%d"),
                            "R$ Preço": st.column_config.NumberColumn("R$ Preço", format="R$ %.2f", disabled=False),
                            "R$ Total": st.column_config.NumberColumn("R$ Total", format="R$ %.2f", disabled=True)
                        }
                        
                        df_forn_edit = st.data_editor(
                            df_exibicao, 
                            hide_index=True, 
                            use_container_width=True, 
                            column_config=col_cfg_forn,
                            height=altura_dinamica,
                            num_rows="fixed",
                            key=f"forn_{fornecedor}_{st.session_state['reset_counter']}"
                        )
                        
                        soma_dinamica = (pd.to_numeric(df_forn_edit["Total"], errors='coerce').fillna(0) * pd.to_numeric(df_forn_edit["R$ Preço"], errors='coerce').fillna(0)).sum()
                        
                        st.markdown(f"""
                            <div style="text-align:right; font-weight:700; margin-top:8px; color:var(--orange-bright); font-size:16px;">
                                Total Final: R$ {soma_dinamica:,.2f}
                            </div>
                        """, unsafe_allow_html=True)
        st.write("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# ROTA 5: CATÁLOGO DE PRODUTOS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Catálogo de Produtos":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--orange-dark) 0%, #241005 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🏷️</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Catálogo de Produtos</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Vincule o Cód. Prime do ERP, gerencie itens e defina o setor e as permissões por loja.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        st.caption("➕ Adicione produtos na última linha  •  🗑️ Selecione a linha e pressione **Delete** para remover  •  ✅ Checkboxes controlam visibilidade por loja")
        
        # --- BOTOES ADICIONADOS AQUI ---
        col_btn1, col_btn2, col_info = st.columns([2.5, 2.5, 5])
        with col_btn1:
            btn_salvar = st.button("💾 Salvar Códigos e Catálogo", type="primary", use_container_width=True)
        with col_btn2:
            btn_atualizar_90d = st.button("🔄 Atualizar Média 90d", use_container_width=True)
        with col_info:
            st.info("💡 Digite o Cód. ERP e salve, ou clique para extrair a Média de Vendas do ERP!")

        # --- LÓGICA DE EXECUÇÃO DA QUERY ---
        if btn_atualizar_90d:
            with st.spinner("⏳ Extraindo vendas dos últimos 90 dias do ERP. Isso pode demorar alguns segundos..."):
                try:
                    conn_pg = st.connection("banco_erp", type="sql")
                    
                    query_90d = 'SELECT loja, codigo, qtde FROM "python_90dSEGTER"'
                    
                    df_nova_media = conn_pg.query(query_90d, ttl=0) # Força rodar na hora
                    
                    # Salva no Sheets
                    conn.update(worksheet=WS_MEDIA_90, data=df_nova_media)
                    st.cache_data.clear() # Limpa cache pro sistema puxar da planilha
                    
                    st.success("✅ Média de 90 dias calculada e salva no Google Sheets com sucesso!")
                except Exception as e:
                    st.error(f"⚠️ Ocorreu um erro ao atualizar os dados no ERP: {e}")

        df_cat_edit = df_produtos.copy()
        df_cat_edit["Cód.Prime"] = df_cat_edit["Cód.Prime"].replace(0, None)

        config_catalogo = {
            "Cód.Prime": st.column_config.NumberColumn("Cód. ERP (Prime)", width=120, min_value=0, format="%d"),
            "Código":    st.column_config.NumberColumn("Cód. Iceasa", width=90, required=True, min_value=0, format="%d"),
            "Descrição": st.column_config.TextColumn("Descrição do Item", width=310, required=True),
            "Tipo":      st.column_config.SelectboxColumn("Setor", options=["Box", "Pedra"], width=100, required=True),
        }
        for loja in LOJAS:
            config_catalogo[loja] = st.column_config.CheckboxColumn(loja, default=True, width=70)

        df_cat_editado = st.data_editor(
            df_cat_edit,
            num_rows="dynamic",
            column_config=config_catalogo,
            hide_index=True,
            use_container_width=True,
            height=580
        )

        if btn_salvar:
            df_cat_editado["Cód.Prime"] = pd.to_numeric(df_cat_editado["Cód.Prime"], errors='coerce').fillna(0).astype(int)
            conn.update(worksheet=WS_PRODUTOS, data=df_cat_editado)
            st.cache_data.clear()
            st.success("✅ Catálogo e permissões atualizados para todas as lojas!")
            st.rerun()
